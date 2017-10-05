import csv
import json
import pdb
import sqlite3
from random import randrange
from pprint import pprint
from openpyxl import load_workbook

class PickStudent:
	already_picked = []

	def reset(self):
		self.already_picked = []

	def save_picks(self, file):
		with open(file, 'w', newline='') as csvfile:
			writer = csv.writer(csvfile)
			writer.writerows(self.already_picked)

	def get_students(self, file):
		'''
		Veja https://docs.python.org/3/library/csv.html
		'''
		with open(file, newline='') as csvfile:
			return list(csv.reader(csvfile))

	def get_students_json(self, file):
		'''
		Veja https://docs.python.org/3/library/json.html
		'''
		with open(file, newline='') as jsonfile:
			data = json.load(jsonfile)
			return data["alunos"]


	def get_students_sqlite(self, file):
		data = []
		conn = sqlite3.connect(file)
		query = '''select * from alunos'''
		cursor = conn.execute(query)
		for index, row in enumerate(cursor): data.append(list(row))
		conn.close()
		return data

	def get_students_xlsx(self, file):
		'''
		Veja https://openpyxl.readthedocs.io/en/default/tutorial.html#loading-from-a-file
		E http://zetcode.com/articles/openpyxl/
		'''
		wb = load_workbook(file)
		sheet = wb.active
		row = 1
		data = []
		while row<1000:
			number = str(sheet.cell(row=row, column=1).value)
			name = sheet.cell(row=row, column=2).value
			row += 1
			if name:
				data.append([number, name])
			else:
				break

		return data

	def pick_random(self, list):
		which = randrange(len(list))
		pick = list[which]
		return pick

	def pick_one(self, list):
		# pprint(self.already_picked)
		pick = self.pick_random(list)
		max_tries = len(list)*100	# prevents running forever; unlikely but it could happen since it's random
		while pick in self.already_picked and max_tries > 0:
			# pprint(f"Ops! Already picked {pick}")
			pick = self.pick_random(list)
			max_tries -= 1

		if max_tries > 0:
			# pprint(f"Aha! Picked {pick}")
			self.already_picked.append(pick)
			return pick
		else:
			raise IndexError




if __name__ == '__main__':
	ps = PickStudent()
	alunos = ps.get_students('./alunos.csv')
	user_input = None
	count = 1
	while not user_input and count < len(alunos):
		print(ps.pick_one(alunos)[1])
		count += 1
		user_input = input("Next...")
	ps.save_picks('./selected.csv')
